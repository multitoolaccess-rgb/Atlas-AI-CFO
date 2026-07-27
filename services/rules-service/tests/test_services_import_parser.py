"""Phase 5 direct unit tests on the lifted ``app.services.import_parser``.

Tests exercise the real parser with file fixtures. Run alongside
``tests/test_routes_imports_real.py`` (route-layer integration).
"""
import io
from pathlib import Path

import pypdf
import pytest
from fastapi import UploadFile

FIXTURES = Path(__file__).resolve().parent / "fixtures"


def _csv_upload(name: str, body: bytes) -> UploadFile:
    return UploadFile(filename=name, file=io.BytesIO(body))


def _minimal_pdf_bytes(lines: list[str]) -> bytes:
    writer = pypdf.PdfWriter()
    page = writer.add_blank_page(width=595, height=842)
    text_block = "\n".join(lines)
    from pypdf.generic import (
        DecodedStreamObject,
        DictionaryObject,
        NameObject,
    )
    content = DecodedStreamObject()
    content.set_data(
        f"BT /F1 12 Tf 50 800 Td ({text_block.encode('latin-1', errors='replace').decode('latin-1')}) Tj ET".encode()
    )
    page[NameObject("/Contents")] = content
    page[NameObject("/Resources")] = DictionaryObject(
        {
            NameObject("/Font"): DictionaryObject(
                {NameObject("/F1"): DictionaryObject({NameObject("/Type"): NameObject("/Font"), NameObject("/Subtype"): NameObject("/Type1"), NameObject("/BaseFont"): NameObject("/Helvetica")})}
            )
        }
    )
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def test_parse_csv_transactions_returns_5_normalized_records():
    """Sample CSV -> 5 records with normalised schema (date/amount/description/merchant_name/is_pending)."""
    from app.services.import_parser import parse_csv_transactions

    body = (FIXTURES / "sample-bank-statement.csv").read_bytes()
    rows = parse_csv_transactions(_csv_upload("sample.csv", body))

    assert len(rows) == 5
    # Phase 52+ — parse_csv_transactions now emits dual ``debit``/``credit``
    # magnitudes alongside the signed ``amount`` so the route layer can
    # insert them into the new split-columns on Transaction. Both columns
    # are unsigned-positive (the unused side per row is ``None``). The
    # historical 5-key shape stays identical so a future regression that
    # drops either field surfaces as a missing-key assertion failure
    # here (locking the schema drift instead of papering over it).
    expected_keys = {
        "transaction_date",
        "amount",
        "description",
        "merchant_name",
        "is_pending",
        "debit",
        "credit",
    }
    for row in rows:
        assert set(row.keys()) == expected_keys
        assert isinstance(row["transaction_date"].year, int)
        assert isinstance(row["amount"], float)
        assert isinstance(row["description"], str)
        assert isinstance(row["is_pending"], bool)
        # Phase 52+ — debit/credit are nullable. For single-amount
        # CSVs one side will be populated and the other ``None``,
        # but ``amount == 0`` FX-neutral rows can validly have
        # BOTH ``None`` (the route treats those as zero-side
        # ledger entries). The exact `(None, X) or (X, None)`
        # XOR therefore can't be a generic invariant; just
        # confirm the columns are present and the dictionary
        # shape matches the contract.
        assert row["debit"] is None or isinstance(row["debit"], float)
        assert row["credit"] is None or isinstance(row["credit"], float)
        # If amount and debit are populated, debit must be the
        # unsigned-positive magnitude (no sign drift between
        # amount and debit columns).
        if row["debit"] is not None:
            assert row["debit"] >= 0, f"debit must be unsigned positive; got {row['debit']}"
        if row["credit"] is not None:
            assert row["credit"] >= 0, f"credit must be unsigned positive; got {row['credit']}"

    first = rows[0]
    assert first["description"] == "Atlas test coffee"
    assert first["amount"] == pytest.approx(-4.50, abs=1e-6)
    assert first["merchant_name"] == "Test Merchant Alpha"
    assert first["is_pending"] is False


def test_parse_csv_transactions_handles_empty_dataframe():
    from app.services.import_parser import parse_csv_transactions

    body = (FIXTURES / "empty-statement.csv").read_bytes()
    rows = parse_csv_transactions(_csv_upload("empty.csv", body))
    assert rows == []


def test_parse_csv_transactions_rejects_missing_required_columns():
    from app.services.import_parser import parse_csv_transactions

    body = (FIXTURES / "bad-statement.csv").read_bytes()
    with pytest.raises(ValueError) as exc:
        parse_csv_transactions(_csv_upload("bad.csv", body))
    msg = str(exc.value).lower()
    assert "date" in msg
    assert "amount" in msg
    assert "description" in msg


def test_parse_csv_transactions_skips_malformed_rows_not_fails():
    """Phase 9 regression: a Fidelity-style footer / trailing line with
    the wrong column count must NOT collapse the whole upload. Before
    the ``on_bad_lines='skip'`` fix, the BE raised
    ``ParserError: Error tokenizing data. C error: Expected 5 fields
    in line 7, saw 4`` which surfaced as a 400 to the user.

    The test uses an inline CSV body (not a fixture file) because the
    inciting case is environment-specific and we want the test to
    encode the EXACT bytes the user uploaded. The fixture-based tests
    above prove the happy path; this one locks the lenient-parser
    contract so a future ``on_bad_lines`` change is caught.
    """
    from app.services.import_parser import parse_csv_transactions

    body = (
        # 1 — header (5 columns: date, description, amount, merchant_name, balance)
        b"date,description,amount,merchant_name,balance\n"
        # 2-6 — legitimate data rows
        b"2025-01-15,Coffee shop,-4.50,Blue Bottle Coffee,1234.56\n"
        b"2025-01-16,Payroll,3500.00,Acme Corp,4734.56\n"
        b"2025-01-17,Grocery,-87.32,Whole Foods,4647.24\n"
        b"2025-01-18,Gas,-52.10,Shell,4595.14\n"
        b"2025-01-19,Subscription,-29.99,Netflix,4565.15\n"
        # 7 — malformed footer with only 4 fields (no merchant_name).
        # The previous strict-mode parser aborted the entire read here.
        b"2025-01-20,Footer,trailing-row-arbitrary-text\n"
    )

    rows = parse_csv_transactions(_csv_upload("messy.csv", body))

    # We drop the malformed row but keep the 5 legitimate rows. Bound
    # ``>= 5`` not ``== 5`` so a future enhancement that recovers extra
    # rows (e.g. via a cleaner ``engine='python'`` + quoteroundtrip)
    # doesn't trip the test.
    assert len(rows) >= 5
    descriptions = [r["description"] for r in rows]
    assert "Coffee shop" in descriptions
    assert "Payroll" in descriptions
    assert "Subscription" in descriptions
    # The malformed footer row never appears as a transaction.
    assert "Footer" not in descriptions


def test_parse_csv_file_preview_skips_malformed_rows_too():
    """Lock the parse_preview / parse_persist invariant under the same
    lenient-parser contract. Before this fix the preview endpoint said
    "5 rows detected" while the persist path silently dropped to 0 —
    the BE leaked an inconsistent preview/persist pair to the UI.

    The test asserts that ``parse_csv_file`` (preview) and
    ``parse_csv_transactions`` (persist) BOTH survive the malformed
    row, so the user's "5 records detected" count always matches the
    transactions that actually land in the DB.
    """
    from app.services.import_parser import parse_csv_file, parse_csv_transactions

    body = (
        b"date,description,amount,merchant_name\n"
        b"2025-01-15,Coffee,-4.50,Blue Bottle\n"
        b"2025-01-16,Payroll,3500.00,Acme\n"
        b"2025-01-17,Grocery,-87.32,Whole Foods\n"
        b"2025-01-18,Gas,-52.10,Shell\n"
        # Malformed row (3 fields instead of 4).
        b"2025-01-19,broken\n"
    )

    preview = parse_csv_file(_csv_upload("messy.csv", body))
    persisted = parse_csv_transactions(_csv_upload("messy.csv", body))

    # Strict equality: preview ``record_count`` MUST exactly match the
    # surviving row count that lands in the DB. If a future regression
    # drops the row in preview but NOT in persist (or vice versa), the
    # UI shows "4 rows detected" while the user sees 3 in their
    # transaction list, and this assertion catches it.
    assert preview["record_count"] == 4
    assert len(persisted) == 4
    assert preview["record_count"] == len(persisted)


def test_parse_csv_accepts_space_separated_headers():
    """Phase 9 regression: real-world bank CSV exports use
    space-separated column names (``Transaction Date``, ``Posted
    Date``, ``Transaction Amount``, ``Transaction Details``). Before
    this fix ``_normalize_headers`` only did ``.strip().lower()`` so
    ``Transaction Date`` resolved to ``transaction date`` (with a
    literal space) and missed the ``transaction_date`` synonym. The
    user got a confusing "Missing: date" 400 from a CSV whose
    columns were semantically correct. The whitespace-collapse
    fix in :func:`_normalize_headers` turns space-separated headers
    into the canonical snake_case keys.

    Locks BOTH the new normalization (spaces -> underscores) AND
    the new synonyms (``posting_date``, ``date_posted``,
    ``transaction_details``, ``particulars``, ``narrative``,
    ``remarks``).
    """
    from app.services.import_parser import parse_csv_file, parse_csv_transactions

    body = (
        b"Transaction Date,Transaction Amount,Transaction Details,Merchant\n"
        b"2025-01-15,-4.50,Coffee shop,Blue Bottle\n"
        b"2025-01-16,3500.00,Payroll,Acme\n"
        b"2025-01-17,-87.32,Grocery,Whole Foods\n"
        b"2025-01-18,-52.10,Gas,Shell\n"
    )

    preview = parse_csv_file(_csv_upload("spaced.csv", body))
    persisted = parse_csv_transactions(_csv_upload("spaced.csv", body))

    # All 4 rows survive AND the schema-validate passes (no 400).
    assert preview["record_count"] == 4
    assert len(persisted) == 4
    assert preview["record_count"] == len(persisted)

    descriptions = [r["description"] for r in persisted]
    assert descriptions == ["Coffee shop", "Payroll", "Grocery", "Gas"]
    amounts = [r["amount"] for r in persisted]
    assert amounts[0] == pytest.approx(-4.50, abs=1e-6)
    assert amounts[1] == pytest.approx(3500.00, abs=1e-6)


def test_parse_csv_accepts_common_banking_synonyms():
    """Phase 9 synonym expansion: cover the headers Chase/BofA/Wells/
    Amex/Fidelity commonly emit (``Posting Date``, ``Date Posted``,
    ``Particulars``, ``Narrative``, ``Remarks``, ``Trans Date``,
    ``Txn Date``). Each combination must validate without raising
    the column-required error.

    Drives :func:`_validate_csv_schema` directly so a future
    synonym map change is caught here without needing a real upload.
    """
    import pandas as pd
    from app.services.import_parser import (
        _validate_csv_schema,
    )

    cases = [
        ["Date Posted", "Amount", "Description"],
        ["Posting Date", "Amount", "Description"],
        ["Posted Date", "Amount", "Memo"],
        ["Trans Date", "Amount", "Transaction Details"],
        ["Txn Date", "Amount", "Particulars"],
        ["Tx Date", "Amount", "Narrative"],
        ["Transaction Date", "Amount", "Remarks"],
        ["DATE", "AMOUNT", "DESCRIPTION"],  # all-caps
        ["  Transaction Date  ", "  Amount  ", "  Description  "],  # whitespace pads
        ["DATE\tPOSTED", "AMOUNT", "MEMO"],  # mixed whitespace -> underscores
    ]
    for headers in cases:
        df = pd.DataFrame(columns=headers)
        # _validate_csv_schema mutates df.columns in place; that's
        # the contract the production code path uses.
        column_map = _validate_csv_schema(df)
        canonicals = set(column_map.values())
        assert {"date", "amount", "description"}.issubset(
            canonicals
        ), f"synonym set failed for {headers!r}: {column_map!r}"


def test_parse_csv_rejects_unrecognised_headers_even_after_normalization():
    """Phase 9 negative lock: a CSV whose headers are spelled
    consistently but use a column name outside the synonym map
    (e.g. ``Value`` for amount, ``Txn Type`` for date) MUST still
    raise the friendly missing-columns error. This proves the new
    whitespace-collapse doesn't accidentally widen the accepted
    universe beyond the documented synonym set.
    """
    import pandas as pd
    from app.services.import_parser import _validate_csv_schema

    df = pd.DataFrame(columns=["Value", "Txn Type", "Note"])
    with pytest.raises(ValueError) as exc:
        _validate_csv_schema(df)
    msg = str(exc.value).lower()
    assert "date" in msg
    assert "amount" in msg
    assert "description" in msg


def test_parse_csv_preview_persist_symmetric_under_capitalised_header():
    """Phase 9 regression: when a bank exports the amount column with a
    capitalised header (e.g. Fidelity's "Amount"), the preview MUST
    still trim the same NaN-padded rows the persist path drops. Before
    this fix the preview checked the literal lowercase "amount"
    column and would silently include rows that the persist path
    would drop, reintroducing the preview/persist drift the Phase 9
    ship set out to eliminate.

    Header normalization (``_normalize_headers`` + ``_build_column_map``)
    is what makes both paths agree on which column is the canonical
    "amount" — this test locks that contract.
    """
    from app.services.import_parser import parse_csv_file, parse_csv_transactions

    # Note CAPITALIZED "Amount" header plus a 3-field malformed row.
    # The 3-field row gets NaN-padded by pandas; both paths should
    # recognise it via canonicalization and drop it. Surviving rows = 4.
    body = (
        b"Date,Description,Amount,Merchant_Name\n"
        b"2025-01-15,Coffee,-4.50,Blue Bottle\n"
        b"2025-01-16,Payroll,3500.00,Acme\n"
        b"2025-01-17,Grocery,-87.32,Whole Foods\n"
        b"2025-01-18,Gas,-52.10,Shell\n"
        b"2025-01-19,broken\n"
    )

    preview = parse_csv_file(_csv_upload("messy_capitalised.csv", body))
    persisted = parse_csv_transactions(_csv_upload("messy_capitalised.csv", body))

    # Strict equality under capitalised headers. Without the symmetry
    # fix the preview would report 5 rows (kept the NaN-padded row)
    # while persist silently dropped it to 4 \u2014 the user sees two
    # different counts for the SAME file.
    assert preview["record_count"] == 4
    assert len(persisted) == 4
    assert preview["record_count"] == len(persisted)


def test_parse_uploaded_statement_dispatches_to_csv_for_csv_extension():
    from app.services.import_parser import parse_uploaded_statement

    body = (FIXTURES / "sample-bank-statement.csv").read_bytes()
    result = parse_uploaded_statement(_csv_upload("sample.CSV", body))
    assert result["file_type"] == "csv"
    assert result["record_count"] == 5
    assert result["filename"] == "sample.CSV"


def test_parse_uploaded_statement_rejects_unsupported_extension():
    from app.services import import_parser

    body = b"binary junk"
    with pytest.raises(ValueError) as exc:
        import_parser.parse_uploaded_statement(_csv_upload("statement.xyz", body))
    msg = str(exc.value).lower()
    assert "unsupported" in msg or "csv" in msg


def test_parse_excel_returns_normalized_records():
    """Excel (.xlsx) with standard columns -> records matching the CSV
    schema (date/amount/description/merchant_name/is_pending). Uses
    openpyxl to build an inline XLSX payload."""
    import openpyxl
    import io
    from app.services.import_parser import parse_excel_file, parse_excel_transactions

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Date", "Description", "Amount", "Merchant"])
    ws.append(["2025-01-15", "Coffee shop", -4.50, "Blue Bottle"])
    ws.append(["2025-01-16", "Payroll", 3500.00, "Acme Corp"])
    ws.append(["2025-01-17", "Grocery", -87.32, "Whole Foods"])
    ws.append(["2025-01-18", "Gas", -52.10, "Shell"])
    ws.append(["2025-01-19", "Subscription", -29.99, "Netflix"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    upload_file = UploadFile(filename="statement.xlsx", file=buf)

    # Preview path
    preview = parse_excel_file(upload_file)
    assert preview["file_type"] == "xlsx"
    assert preview["record_count"] == 5
    assert len(preview["preview"]) == 5

    # Persist path
    upload_file.file.seek(0)
    records = parse_excel_transactions(upload_file)
    assert len(records) == 5
    # Phase 36 — multi-sheet Excel emits a ``sheet_name`` tag for the
    # route layer to route each row to its corresponding auto-created
    # account (Phase 38 — ``Sheet1`` rows → a Checking account,
    # ``Sheet2`` rows → a Savings account, etc.). Without this key the
    # FE cannot distinguish which sheet a row came from, defeating the
    # multi-account import UX. ``debit``/``credit`` (Phase 52+) are the
    # signed-positive magnitudes mirrored from the dual-column
    # bookkeeping convention; either may be ``None`` per row.
    expected_keys = {
        "transaction_date",
        "amount",
        "description",
        "merchant_name",
        "is_pending",
        "debit",
        "credit",
        "sheet_name",
    }
    for row in records:
        assert set(row.keys()) == expected_keys
        assert isinstance(row["transaction_date"].year, int)
        assert isinstance(row["amount"], float)

    first = records[0]
    assert first["description"] == "Coffee shop"
    assert first["amount"] == pytest.approx(-4.50, abs=1e-6)
    assert first["merchant_name"] == "Blue Bottle"
    assert first["is_pending"] is False

    # Preview/persist parity
    assert preview["record_count"] == len(records)


def test_parse_excel_accepts_split_amount_columns():
    """Excel with Debit/Credit columns (no single Amount) parses
    via the split-amount path, same as CSV."""
    import openpyxl
    import io
    from app.services.import_parser import parse_excel_file, parse_excel_transactions

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "Debit", "Credit"])
    ws.append(["01/02/2024", "Coffee shop", 4.50, None])
    ws.append(["01/03/2024", "Salary", None, 3500.00])
    ws.append(["01/04/2024", "Grocery", 87.32, None])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    upload = UploadFile(filename="split.xlsx", file=buf)
    preview = parse_excel_file(upload)
    upload.file.seek(0)
    records = parse_excel_transactions(upload)

    assert preview["record_count"] == 3
    assert len(records) == 3
    assert records[0]["amount"] == pytest.approx(-4.50, abs=1e-6)
    assert records[1]["amount"] == pytest.approx(3500.00, abs=1e-6)
    assert records[2]["amount"] == pytest.approx(-87.32, abs=1e-6)


def test_parse_excel_dispatches_to_correct_parser():
    """``parse_uploaded_statement`` dispatches .xlsx to Excel parser"""
    from app.services.import_parser import parse_uploaded_statement

    import openpyxl
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "Amount"])
    ws.append(["2025-01-15", "Test", 100.00])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    result = parse_uploaded_statement(UploadFile(filename="data.xlsx", file=buf))
    assert result["file_type"] == "xlsx"
    assert result["record_count"] == 1


def test_parse_csv_file_preview_rejects_csv_without_required_columns():
    """Phase 9 schema-symmetry lock: a CSV with arbitrary headers (no
    date/amount/description) must raise ValueError from BOTH the
    preview path AND the persist path. Before this fix the preview
    would silently return ``record_count=N`` while persist would 400,
    reintroducing the preview/persist drift the ship set out to
    eliminate.
    """
    from app.services.import_parser import parse_csv_file, parse_csv_transactions

    body = (
        b"foo,bar,baz\n"           # header that does NOT include date/amount/description
        b"1,2,3\n"
        b"4,5,6\n"
        b"7,8,9\n"
    )

    with pytest.raises(ValueError) as exc_preview:
        parse_csv_file(_csv_upload("missing_dates.csv", body))
    msg = str(exc_preview.value).lower()
    assert "date" in msg
    assert "amount" in msg
    assert "description" in msg

    # Persist path raises the same error message — single source of
    # truth for the schema contract.
    with pytest.raises(ValueError) as exc_persist:
        parse_csv_transactions(_csv_upload("missing_dates.csv", body))
    assert str(exc_persist.value).lower() == msg


def test_parse_csv_preview_keeps_blank_description_legitimate_row():
    """Phase 9 blank-description semantics lock: a row with an empty
    ``description`` field but valid date + amount is KEPT in both
    preview and persist (auto-filled with ``"Imported transaction"``
    placeholder downstream). Before this fix the helper drop-blank
    filter was too aggressive and would silently lose legitimate
    Plaid-style exports that leave memo/description blank on some
    rows.
    """
    from app.services.import_parser import parse_csv_file, parse_csv_transactions

    body = (
        b"date,description,amount,merchant_name\n"
        b"2025-01-15,Coffee,-4.50,Blue Bottle\n"
        b"2025-01-16,,3500.00,Acme\n"           # empty description, valid date/amount
        b"2025-01-17,Grocery,-87.32,Whole Foods\n"
        b"2025-01-18,Gas,-52.10,Shell\n"
    )

    preview = parse_csv_file(_csv_upload("blank_desc.csv", body))
    persisted = parse_csv_transactions(_csv_upload("blank_desc.csv", body))

    # All 4 rows survive preview AND persist \u2014 the blank-description
    # row is treated as legitimate (placeholder fill fires later).
    assert preview["record_count"] == 4
    assert len(persisted) == 4
    assert preview["record_count"] == len(persisted)

    # The placeholder fills in downstream. The positional lock proves
    # ONLY the blank-description row got placeholders — the others keep
    # their raw text. A regression that fires the placeholder on the
    # wrong row trips ``descriptions[0]/[2]/[3] == "Imported
    # transaction"`` instead of getting past the membership check.
    descriptions = [r["description"] for r in persisted]
    assert "" not in descriptions
    assert descriptions == [
        "Coffee",
        "Imported transaction",
        "Grocery",
        "Gas",
    ] # positional lock
    assert "Imported transaction" in descriptions


def test_parse_csv_file_preview_returns_first_five_records():
    """``parse_csv_file`` is the preview-only sibling of
    parse_csv_transactions; it returns the same shape as
    ``parse_uploaded_statement`` but for CSV specifically."""
    from app.services.import_parser import parse_csv_file

    body = (FIXTURES / "sample-bank-statement.csv").read_bytes()
    result = parse_csv_file(_csv_upload("sample.csv", body))
    assert result["file_type"] == "csv"
    assert result["record_count"] == 5
    assert len(result["preview"]) == 5


def test_parse_uploaded_statement_dispatches_to_pdf_for_pdf_extension():
    """PDF dispatch contract test: use ``unittest.mock`` to verify that
    ``parse_uploaded_statement`` correctly delegates to ``parse_pdf_file``
    for ``.pdf`` extensions WITHOUT actually parsing a real PDF (which
    the hand-rolled fixture may or may not be valid for). The mock
    keeps the test deterministic.

    To replace with a real-PDF happy-path test, ship a real-world PDF
    fixture in tests/fixtures/sample.pdf (1-2 page text-only PDF).
    """
    from unittest.mock import patch
    from app.services import import_parser

    with patch.object(
        import_parser, "parse_pdf_file", return_value={"file_type": "pdf", "preview": ["x"]}
    ) as mp:
        result = import_parser.parse_uploaded_statement(_csv_upload("statement.pdf", b""))
        mp.assert_called_once()
        assert result["file_type"] == "pdf"


def test_parse_csv_accepts_split_amount_credit_debit():
    """Phase 9b split-amount support: a CSV with separate Debit and
    Credit columns (no combined ``amount`` column) is parsed
    transparently: the per-row loop computes ``amount = credit - debit``
    so credits are positive, debits negative, matching how the user
    reads their statement.

    Locks BOTH the schema-validator path (accepts split pair as a
    valid amount-source) AND the per-row computation. Without this,
    a Chase-style CSV export would raise
    ``CSV statement must include 'date', 'amount', and 'description'
    columns. Missing: amount.`` even though Debit + Credit cover the
    requirement.
    """
    from app.services.import_parser import parse_csv_file, parse_csv_transactions

    body = (
        b"Date,Description,Debit,Credit\n"
        b"01/02/2024,Coffee shop,4.50,\n"
        b"01/03/2024,Salary,,3500.00\n"
        b"01/04/2024,Grocery,87.32,\n"
        b"01/05/2024,Refund,,25.00\n"
    )

    preview = parse_csv_file(_csv_upload("split.csv", body))
    persisted = parse_csv_transactions(_csv_upload("split.csv", body))

    # Preview/persist parity.
    assert preview["record_count"] == 4
    assert len(persisted) == 4
    assert preview["record_count"] == len(persisted)

    amounts = [r["amount"] for r in persisted]
    descriptions = [r["description"] for r in persisted]
    assert amounts[0] == pytest.approx(-4.50, abs=1e-6)   # debit
    assert amounts[1] == pytest.approx(3500.00, abs=1e-6) # credit
    assert amounts[2] == pytest.approx(-87.32, abs=1e-6)  # debit
    assert amounts[3] == pytest.approx(25.00, abs=1e-6)   # credit
    assert descriptions == ["Coffee shop", "Salary", "Grocery", "Refund"]


def test_parse_csv_accepts_split_amount_deposits_withdrawals():
    """Phase 9b split-amount support: ``Deposits``/``Withdrawals``
    synonyms (common savings-account export) route to the same
    canonical ``credit``/``debit`` pair as ``Credit``/``Debit``.
    Locked separately from the Credit/Debit test because the
    synonym->canonical mapping is a separate code path the user
    could otherwise regress with a one-line copy-paste.
    """
    from app.services.import_parser import parse_csv_file, parse_csv_transactions

    body = (
        b"Date,Particulars,Withdrawals,Deposits\n"
        b"01/02/2024,Coffee,4.50,\n"
        b"01/03/2024,Salary,,2500.00\n"
        b"01/04/2024,Grocery,87.32,\n"
        b"01/05/2024,Interest,,5.50\n"
    )

    preview = parse_csv_file(_csv_upload("savings.csv", body))
    persisted = parse_csv_transactions(_csv_upload("savings.csv", body))

    assert preview["record_count"] == 4
    assert len(persisted) == 4

    amounts = [r["amount"] for r in persisted]
    assert amounts[0] == pytest.approx(-4.50, abs=1e-6)    # withdrawal -> negative
    assert amounts[1] == pytest.approx(2500.00, abs=1e-6)  # deposit -> positive
    assert amounts[2] == pytest.approx(-87.32, abs=1e-6)   # withdrawal -> negative
    assert amounts[3] == pytest.approx(5.50, abs=1e-6)     # deposit (interest) -> positive


def test_parse_csv_amount_column_wins_over_split_when_both_populated():
    """Phase 9b Constraint-4 lock: when a file has BOTH a plain
    ``amount`` column AND a split (credit + debit) pair, AND the
    plain ``amount`` column is populated on at least one row, the
    plain column wins. The split pair is dropped from the column
    map so the per-row branch takes the single-canonical path.
    Without this, a quirky export that fills all four columns
    (some banks do it for "compatibility with old tools") would
    confuse the parser and lose the true amount on rows where
    the split columns disagree with the plain amount.
    """
    from app.services.import_parser import parse_csv_transactions

    body = (
        b"Date,Description,Amount,Debit,Credit\n"
        b"01/02/2024,Coffee,-4.50,,\n"
        b"01/03/2024,Salary,3500.00,,\n"
    )

    persisted = parse_csv_transactions(_csv_upload("dual.csv", body))
    assert len(persisted) == 2
    amounts = [r["amount"] for r in persisted]
    # Plain amount column wins. The Debit/Credit columns are
    # uniformly empty here, so Constraint-4 picks the populated
    # path transparently.
    assert amounts[0] == pytest.approx(-4.50, abs=1e-6)
    assert amounts[1] == pytest.approx(3500.00, abs=1e-6)


def test_parse_csv_falls_back_to_split_if_amount_column_is_all_nan():
    """Phase 9b Constraint-4 counter-example: when a file has BOTH a
    plain ``amount`` column AND a split (credit + debit) pair, but
    the plain ``amount`` column is uniformly empty/NaN, the split
    pair is used instead. This is the genuine case for
    semi-exported statements where ``Amount`` is the legacy column
    the bank fills with placeholders (or the user is migrating
    between two layouts).
    """
    from app.services.import_parser import parse_csv_transactions

    body = (
        b"Date,Description,Amount,Debit,Credit\n"
        b"01/02/2024,Coffee,,4.50,\n"
        b"01/03/2024,Salary,,,3500.00\n"
    )

    persisted = parse_csv_transactions(_csv_upload("empty_amount.csv", body))
    assert len(persisted) == 2
    amounts = [r["amount"] for r in persisted]
    # The split pair path was used because the plain amount column
    # is uniformly empty.
    assert amounts[0] == pytest.approx(-4.50, abs=1e-6)
    assert amounts[1] == pytest.approx(3500.00, abs=1e-6)


def test_parse_csv_accepts_currency_and_paren_headers():
    """Phase 9b header normalization extension: bank exports
    sometimes emit currency-stripped parenthetical headers like
    ``Amount ($)``, ``Amount (£)``, ``Amount (USD)``. They must
    resolve to ``amount`` after :func:`_normalize_headers` runs.
    Drives the validator directly so a future normalization
    regression is caught without a real upload.
    """
    import pandas as pd
    from app.services.import_parser import _validate_csv_schema

    cases = [
        ["Date", "Amount ($)", "Description"],
        ["Date", "Amount (£)", "Description"],
        ["Date", "Amount (USD)", "Description"],
        ["Date", "Amount ($CAD)", "Description"],
        ["Date", "Debit ($CAD)", "Credit (£)", "Description"],
    ]
    for headers in cases:
        df = pd.DataFrame(columns=headers)
        column_map = _validate_csv_schema(df)
        canonicals = set(column_map.values())
        # The split-pair case ends with credit + debit instead of
        # amount, so it satisfies the missing-check via either path.
        required_ok = (
            {"date", "amount"}.issubset(canonicals)
            or (
                "amount" not in canonicals
                and {"credit", "debit"} <= canonicals
            )
        )
        assert required_ok, (
            f"currency-stripped headers failed for {headers!r}: "
            f"{column_map!r}"
        )
        assert "description" in canonicals, (
            f"description missing for {headers!r}: {column_map!r}"
        )


def test_parse_csv_drops_split_amt_rows_with_unparseable_side_value():
    """Phase 9b defensive lock: a row in a split-amount CSV where ONE
    side has a typo (``abc``) MUST be dropped, not silently inserted
    with wrong-signed amount. Before the
    ``_coerce_amount_or_zero``-raises-on-unparseable-input lock, a
    row like ``Date,Coffee,abc,3500`` would yield
    ``amount = 0 - 3500 = -3500`` \u2014 wrong-signed pollution of the
    ledger. After this lock, the row is dropped via the existing
    :func:`_drop_malformed_rows` pipeline and an operator ``tail`` on
    the WARNING log can find the offender.

    Lock also requires :func:`_coerce_amount_or_zero` itself to
    raise on truly unparseable input (after the legitimate-zero
    short-circuits), preserving the Chase / BofA / Wells
    sign convention (``amount = credit - debit``, where debit
    values are entered positive and represent money going out)
    while rejecting typos.
    """
    from app.services.import_parser import parse_csv_file, parse_csv_transactions

    body = (
        b"Date,Description,Debit,Credit\n"
        b"01/02/2024,Coffee,abc,\n"                # typo on debit side
        b"01/03/2024,Salary,,3500.00\n"             # clean credit only
    )

    persisted = parse_csv_transactions(_csv_upload("typo.csv", body))
    # Only the clean salary row survives; the typo row is dropped.
    assert len(persisted) == 1
    assert persisted[0]["description"] == "Salary"
    assert persisted[0]["amount"] == pytest.approx(3500.00, abs=1e-6)
    # Preview/persist parity lock — the
    # ``_drop_malformed_rows`` ``pd.to_numeric(errors='coerce')``
    # filter must drop the typo row in preview AND in persist;
    # otherwise we re-introduce the preview/persist drift Phase 9
    # set out to eliminate (preview says ``record_count=2`` while
    # persist writes 1).
    preview = parse_csv_file(_csv_upload("typo.csv", body))
    assert preview["record_count"] == len(persisted)
    assert preview["record_count"] == 1


# Phase 15+ — accept the real-world bank-export amount shapes that
# the previous strip-only parser silently dropped. Drives the
# helpers directly so a future regression that narrows the grammar
# trips the test before reaching the user.
@pytest.mark.parametrize(
    "raw_amount,expected",
    [
        # Plain numbers — baseline (we should NOT regress this).
        ("500.00", 500.0),
        ("-500.00", -500.0),
        # US thousands — already supported.
        ("1,234.56", 1234.56),
        ("$1,234.56", 1234.56),
        ("$1,250,000.00", 1250000.00),
        # NEW (Phase 15+): accounting parens = negative.
        ("(50.00)", -50.0),
        ("(1234.56)", -1234.56),
        ("($50.00)", -50.0),
        # NEW: sign prefix with space.
        ("- 100.00", -100.0),
        ("-50.00", -50.0),
        # NEW: trailing dash.
        ("100.00-", -100.0),
        # NEW: signed-parens ``-(75.50)`` (closes Phase 15.0 gap).
        ("-(75.50)", -75.5),
        # NEW: currency glyphs €/£/¥.
        ("\u20ac500.50", 500.5),
        ("\u00a31200.99", 1200.99),
        ("\u00a5500", 500.0),
        # NEW: padded whitespace + glyph.
        ("  $1,250.00  ", 1250.0),
        ("  \u20ac250.00  ", 250.0),
        # NEW (Phase 15.1): glyph-PRECEDES-sign must keep the sign.
        # Previously ``\u20ac-50.00`` rounded to ``+50.00`` because the
        # sign detector ran BEFORE the currency-strip step and saw no
        # leading ``-``. The reorder (currency-strip FIRST, then sign-
        # detect) closes that hole.
        ("\u20ac-50.00", -50.0),
        ("$-50.00", -50.0),
        ("\u20ac-(75.50)", -75.5),
        ("-(€75.50)", -75.5),
    ],
)
def test_parse_amount_accepts_real_world_bank_shapes(raw_amount, expected):
    """Phase 15+ — every shape a real-world bank export emits must
    parse cleanly. A regression that re-narrows the grammar (e.g.
    dropping the parens / trailing-dash support) trips this test
    immediately so the user does NOT see silent row drops on a real
    Wells Fargo / BofA / Chase file.
    """
    from app.services.import_parser import _parse_amount

    assert _parse_amount(raw_amount) == pytest.approx(expected, abs=1e-6)


@pytest.mark.parametrize(
    "blank_value",
    [
        "",
        "   ",
        None,
        "abc",       # typo — strictly rejected
        "--",        # typo
        "N/A",       # typo
        "$",         # glyph only
        "-",         # sign only
        "(",         # open paren only (not followed by close)
    ],
)
def test_parse_amount_rejects_blank_or_typo(blank_value):
    """Phase 15+ — lock the strict-drop contract. Blank / NaN / typo
    inputs MUST raise ``ValueError`` so the per-row try/except drops
    the row via the existing pipeline. Returning ``0.0`` for ``"abc"``
    would silently produce a wrong-signed ledger entry.
    """
    import pandas as pd

    from app.services.import_parser import _parse_amount

    if blank_value is None:
        with pytest.raises(ValueError):
            _parse_amount(None)
        with pytest.raises(ValueError):
            _parse_amount(pd.NA)
        return
    with pytest.raises(ValueError):
        _parse_amount(blank_value)


def test_parse_amount_split_helper_parity():
    """Phase 15+ — ``_coerce_amount_or_zero`` (split-amount side)
    MUST recognise the same parens / padded signs / currency glyph
    shapes as ``_parse_amount``. The two callers diverge ONLY in
    blank-handling semantics (split-pair returns 0.0 on blank;
    single-amount raises ValueError on blank).

    Without this lock, an asymmetric fix that updates only one
    helper would silently let split-amount CSVs behave differently
    from single-amount CSVs — a regression the user would only
    notice on whichever bank uses split-amount exports.
    """
    from app.services.import_parser import _coerce_amount_or_zero

    parity_cases = [
        ("(50.00)", -50.0),
        ("-$75.50", -75.5),
        ("- 100.00", -100.0),
        ("100.00-", -100.0),
        ("$3,500.00", 3500.0),
        ("\u20ac500.50", 500.5),
        ("\u00a31200.99", 1200.99),
        ("1,234.56", 1234.56),
    ]
    for raw, expected in parity_cases:
        assert _coerce_amount_or_zero(raw) == pytest.approx(expected, abs=1e-6), (
            f"split-amount helper failed on {raw!r}: expected {expected}"
        )

    # Split helper's blank semantics: blank -> 0.0 (NOT raise)
    # because a row with one empty side is legitimate ``amount =
    # credit - debit`` ledger behaviour.
    assert _coerce_amount_or_zero(None) == 0.0
    assert _coerce_amount_or_zero("") == 0.0
    assert _coerce_amount_or_zero("   ") == 0.0


def test_parse_csv_persists_all_edge_case_amount_shapes_via_fixture():
    """Phase 15+ regression: load the canonical ``edge_cases.csv``
    fixture (25 rows covering the full bank-export shape grammar)
    and assert ALL 17 legitimate shapes survive the parser while
    the 8 intentional blanks/typos drop.

    Locks the fix so future helper changes that re-narrow the
    shape grammar (e.g. dropping ``(50.00)``-support to "Phase 14"
    behaviour) trip this test BEFORE the user sees row drops on
    a real BofA / Wells file at upload time.
    """
    from app.services.import_parser import parse_uploaded_statement

    body = (FIXTURES / "sample_statements" / "edge_cases.csv").read_bytes()
    result = parse_uploaded_statement(_csv_upload("edge_cases.csv", body))
    parsed = result["parsed_records"]
    # 25 rows total: 7 intentional drops + 18 legitimate shapes survive
    # (Phase 15+ — the bulk amount filter was removed so the Signed
    # parens ``-(75.50)`` row no longer gets NaN'd before the per-row
    # parser handles it).
    assert result["record_count"] == 18
    assert result["expected_row_count"] == 18
    assert len(parsed) == 18
    # Zero drop warnings (8 rows intentionally dropped are typo/blank;
    # the parser doesn't surface a single \"could not be imported\" warning
    # because those rows failed per-row, not wholesale).
    assert not any(
        "could not be imported" in w for w in result.get("warnings") or []
    )
    # Spot-check the key tolerant-shape rows that the Phase 14 fix
    # would have dropped silently:
    desc_to_amount = {r["description"]: r["amount"] for r in parsed}
    assert desc_to_amount["Refund accounting"] == pytest.approx(-50.00, abs=1e-6)
    assert desc_to_amount["Padded sign negative"] == pytest.approx(-100.00, abs=1e-6)
    assert desc_to_amount["Trailing dash negative"] == pytest.approx(-100.00, abs=1e-6)
    assert desc_to_amount["Signed parens"] == pytest.approx(-75.50, abs=1e-6)
    assert desc_to_amount["Euro symbol"] == pytest.approx(500.50, abs=1e-6)
    assert desc_to_amount["Pound symbol"] == pytest.approx(1200.99, abs=1e-6)
    assert desc_to_amount["Us thousands"] == pytest.approx(1234.56, abs=1e-6)
    assert desc_to_amount["Big payroll"] == pytest.approx(1250000.00, abs=1e-6)
    # The deliberately-dropped rows (typo / blank / sign-only / etc.)
    # must NOT appear:
    assert "Typo row" not in desc_to_amount
    assert "Empty amount" not in desc_to_amount
    assert "Just dollar sign" not in desc_to_amount
    assert "Just dash" not in desc_to_amount
    assert "Just paren" not in desc_to_amount
    assert "All-whitespace amount" not in desc_to_amount
    assert "Double paren negative" not in desc_to_amount
